from django.shortcuts import render, redirect
from .forms import StudentContactForm
from .models import StudentContact, LeadCallStatus
from django.contrib import messages
import xlrd
from openpyxl import load_workbook
from django.utils.dateparse import parse_date
from datetime import datetime
from django.contrib.auth.models import User

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import date, datetime
from django.contrib.auth.decorators import login_required


from xhtml2pdf import pisa
from datetime import datetime, timedelta
from django.utils import timezone
from django.http import HttpResponse
import csv

# Create your views here.

# all contact tab can view with pagenator no filter for contact 
@login_required(login_url="login")
def AddContact(request):

    contacts = StudentContact.objects.filter(active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)

    form  = StudentContactForm()
    if request.method == "POST":
        form = StudentContactForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Data Added To database")
            return redirect("AddContact")
        else:
            messages.error(request,"Data is not saved review the data...")
        #     return redirect("AddContact")

            

    context = {
        "form":form,
        "contacts":page_obj
    }
    return render(request,"addcontact.html",context)


# Exporting data from an excelsheet on all contact tab

@login_required(login_url="login")
def import_data_from_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        if str(excel_file).split(".")[-1].lower() == 'xls':
            workbook = xlrd.open_workbook(file_contents=excel_file.read())
            worksheet = workbook.sheet_by_index(0)
        else:
            workbook = load_workbook(excel_file)
            worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):

            if row[1] == None or row[2] == None:
                continue
            else:
               
                # StudentContact.objects.create(
                #     name=str(row[1]),
                #     contact_number=int(row[2]),
                #     last_status=c,
                #     collage=str(row[6])
                # )

                print(str(row[5]))
                print(str(row[1]))
                print(str(row[6]))
                print((row[3]))

                try:
                    print(int(row[2]))
                    contact = int(row[2])
                except:
                    continue
                try:
                    name = str(row[1])
                except:
                    continue
                try:
                    last_statu = str(row[4])
                except:
                    last_statu = None 
                try:
                    collage=str(row[5])
                except:
                    collage = None
                try:
                    course=str(row[6])
                except:
                    course = None
                try:
                    no_follow=int(row[7])
                except:
                    no_follow = None
                try:
                    email=int(row[8])
                except:
                    email = None

                if StudentContact.objects.filter(contact_number=contact).exists():
                    continue
                try:
                    contacts = StudentContact.objects.create(name=name,contact_number=contact,last_status=last_statu,study_streem=course,number_follow_up =no_follow,collage=collage,last_follow_up = row[3],email = email)
                    contacts.save()
                except:
                    contacts = StudentContact.objects.create(name=name,contact_number=contact,last_status=last_statu,study_streem=course,collage=collage,email = email)
                    contacts.save()
        messages.info(request,"excel File Updated....")
                
    return redirect("AddContact")


# pending tasks are filterd by follow_up status 
@login_required(login_url="login")
def PendingContacts(request):
    contacts = StudentContact.objects.filter(follow_up_status = "Not Called",active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)


    context = {
        
        "contacts":page_obj
    }
    return render(request,"pendingcontacts.html",context)


# single view contact triggering on verious funtions

@login_required(login_url="login")
def ViewContactData(request,pk):
    contact = StudentContact.objects.get(id = pk)
    leadcall_status = LeadCallStatus.objects.filter(contact = contact)[::-1]
    users = User.objects.all()
    dt = date.today()

    context = {
        "contact":contact,
        "leadcall_status":leadcall_status,
        "users":users,
        "dt":dt
    }
    return render(request,"viewcontactdata.html",context)

#update basic data of contact 

@login_required(login_url="login")
def UpdateBasicData(request,pk):
    contact = StudentContact.objects.get(id = pk)
    if request.method == "POST":
        contact.name = request.POST["name"]
        contact.study_streem = request.POST["streem"]
        contact.collage  = request.POST["collage"]
        contact.email= request.POST["email"]
        contact.save()
        messages.info(request, 'Contact Data Updated...')
        return redirect("ViewContactData",pk = pk)
    return redirect("ViewContactData",pk = pk)

# updating leads status 

@login_required(login_url="login")
def FollowUpDateUpdate(request,pk):
    contact = StudentContact.objects.get(id = pk)
    if request.method == "POST":
        contact.next_follow_up = request.POST["date"]
        contact.save()
        messages.info(request, 'Contact Data Updated...')
        return redirect("ViewContactData",pk = pk)
    return redirect("ViewContactData",pk = pk)

@login_required(login_url="login")
def lead_statusUpdate(request,pk,strval):
    contact = StudentContact.objects.get(id = pk)
    contact.lead_status = strval
    contact.save()
    messages.info(request, 'Contact Data Updated...')
    return redirect("ViewContactData",pk = pk)


@login_required(login_url="login")
def FollowUpUpadte(request,pk):
    contact = StudentContact.objects.get(id = pk)
    from datetime import date as dt
    if request.method == "POST":
        
        date = request.POST["next_follow_up"]
        
        dte = dt.today()
        follow_up_status = request.POST["lead_status"]
        follow_up_comments = request.POST['followupcomment']
        follow_up_by = request.user
        
        try:
            lead = LeadCallStatus.objects.create(contact = contact,follow_up_status = follow_up_status, follow_up_comments = follow_up_comments,follow_up_by  = follow_up_by, next_follow_up = date )
            lead.save()
        except:
            lead = LeadCallStatus.objects.create(contact = contact,follow_up_status = follow_up_status, follow_up_comments = follow_up_comments,follow_up_by  = follow_up_by, next_follow_up = dte )
            lead.save()
            date = dte

        if contact.follow_up_started_date == None: # if the follow up is not started need to update the followup started date
           contact.follow_up_started_date = datetime.now()
           contact.save()
        # lead status updated on the frontend number of followup will increase and alos the last status on Contact table is updated  
        contact.last_follow_up = datetime.now()
        contact.number_follow_up += 1
        contact.follow_up_status = follow_up_status 
        contact.last_status = follow_up_comments
        contact.next_follow_up = date
        contact.lead_follow_up = request.user
        contact.save()
        if contact.follow_up_status == "Rejected" or contact.follow_up_status == "Not intrested" :
            contact.active = False
            contact.next_follow_up = None
            contact.save()
        messages.info(request,"Status updated....")
        
    return redirect("ViewContactData",pk = pk)

# assigning the contact to specific staff

@login_required(login_url="login")
def AssignContacts(request):
    contacts = StudentContact.objects.filter(lead_follow_up = None,active = True)
    contacts_count = StudentContact.objects.filter(lead_follow_up = None,active = True).count()
    users = User.objects.all()
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    # will taken the list forn from end with get list method it has the id of all the contact on str format need to conert it to int and sort with get object method and add user and next_follow_up to the data
    if request.method =="POST":
        selected_contacts = request.POST.getlist('contact_id')
        user_id = request.POST['user']
        date = request.POST['date']
        user = User.objects.get(id = int(user_id))
        print(user,date,"----------------------------------------------")
        print(selected_contacts)
        for item in selected_contacts:
            contact = StudentContact.objects.get(id = int(item))
            contact.lead_follow_up = user 
            contact.next_follow_up = date
            contact.save()

        messages.info(request,"{} Contact assigned to {}".format(len(selected_contacts),user.first_name))
        return redirect("AssignContacts")

    context = {
        "contacts":page_obj,
        "users":users,
        "contacts_count":contacts_count
    }
    return render(request,'assigntostaff.html',context)

@login_required(login_url="login")
def AssignContactsSingle(request,pk):
    contact = StudentContact.objects.get(id = pk)
    if request.method == "POST":
        user_id = request.POST['user']
        user = User.objects.get(id = int(user_id))
        contact.lead_follow_up = user
        contact.save()
        messages.info(request,"New Staff assigned To contact follow up")
        return redirect("ViewContactData",pk = pk)
    return redirect("ViewContactData",pk = pk)
    



@login_required(login_url="login")
def TodaysNewCalls(request):
    contacts= StudentContact.objects.filter(lead_follow_up = request.user,follow_up_status = "Not Called", next_follow_up = date.today(),active = True )
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)

    
    }  
    return render(request,"todaysnewcall.html",context)

@login_required(login_url="login")
def MyAssignments(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,follow_up_status = "Not Called",active = True ).order_by("next_follow_up")
    contacts_count = StudentContact.objects.filter(lead_follow_up = request.user,follow_up_status = "Not Called",active = True).count()
    users = User.objects.all()
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)

    context = {
        "contacts":page_obj,
        "users":users,
        "contacts_count":contacts_count
    }  
    return render(request,"myassignments.html",context)


@login_required(login_url="login")
def PendingTocall(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,next_follow_up__lt = date.today(),active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)
    
    }  

    return render(request,"pendingcontacts.html",context)


@login_required(login_url="login")
def TodaysFollowUp(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,next_follow_up = date.today(),active = True).exclude(follow_up_status = "Not Called")
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)

    
    }  

    return render(request,"todays_follow_up.html",context)

@login_required(login_url="login")
def UpcommingFollowUp(request):
    # contacts = StudentContact.objects.filter(lead_follow_up = request.user,next_follow_up__gt = date.today(),active = True)
    from datetime import date, timedelta

    # Define the date range
    start_date = date.today()
    end_date = start_date + timedelta(days=7)

    # Query the StudentContact objects
    contacts = StudentContact.objects.filter(
        lead_follow_up=request.user,
        next_follow_up__range=(start_date, end_date),
        active=True
    )

    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)

    
    }  

    return render(request,"upcommingfollowup.html",context)


@login_required(login_url="login")
def WarmLeadsUser(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,lead_status ='Warm Lead',active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)

    
    }  
    return render(request,"warmleadsuser.html",context)

@login_required(login_url="login")
def HotLeadsUser(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,lead_status ='Hot Lead',active = True )
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)

    
    }  
    return render(request,"hotleadsuser.html",context)

@login_required(login_url="login")
def ConvertedLeads(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,lead_status ='Converted',active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)

    
    }  
    return render(request,'convertedleads.html',context)

@login_required(login_url="login")
def CompletedToday(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,last_follow_up = date.today())
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts)
    }  

    return render(request,"leadscompletedtoday.html",context)


        
    
# functions for admin Views........................



@login_required(login_url="login")
def UserWiseData(request):
    users = User.objects.all()
    context = {
        "users":users
    }
    return render(request,"userwisedata.html",context)

@login_required(login_url="login")
def UpdatesOfstaff(request,pk):
    today = timezone.now()
    start_date = today + timedelta(days=-1)
    user = User.objects.get(id = pk)
    contact_count = StudentContact.objects.filter(lead_follow_up = user,follow_up_status = "Not Called",active = True ).count()
    new_contact_count = StudentContact.objects.filter(lead_follow_up = user,follow_up_status = "Not Called", next_follow_up = date.today(),active = True ).count()
    penidng_call_list = StudentContact.objects.filter(lead_follow_up = user,next_follow_up__lt = date.today(),active = True).count()
    today_follow_up = StudentContact.objects.filter(lead_follow_up = user,next_follow_up = date.today(),active = True).exclude(follow_up_status = "Not Called").count()
    upcomming_contacts_count = StudentContact.objects.filter(lead_follow_up = user,next_follow_up__gt = date.today(),active = True).count()
    today_contacts_completed = StudentContact.objects.filter(lead_follow_up = user,last_follow_up = date.today(),active = True).count()
    rejected_contacts = StudentContact.objects.filter(lead_follow_up = user, active=False).count()
    yesterday_contacts = StudentContact.objects.filter(lead_follow_up = user,last_follow_up = start_date ).count()
    allcontact = StudentContact.objects.filter(lead_follow_up = user).count()


    context = {
        "contact_count":contact_count,
        "penidng_call_list":penidng_call_list,
        "today_follow_up":today_follow_up,
        "new_contact_count":new_contact_count,
        "upcomming_contacts_count":upcomming_contacts_count,
        "today_contacts_completed":today_contacts_completed,
        "rejected_contacts":rejected_contacts,
        "yesterday_contacts":yesterday_contacts,
        "allcontact":allcontact,
        "user":user
    }

    return render(request,"updatesofstaffs.html",context)


@login_required(login_url="login")
def TodaysNewCallsAdmin(request,pk):
    user1 = User.objects.get(id = pk)
    contacts= StudentContact.objects.filter(lead_follow_up = user1,follow_up_status = "Not Called", next_follow_up = date.today(),active = True )
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts),
    "user1":user1
    }  
    return render(request,"todaysnewcall.html",context)


@login_required(login_url="login")
def PendingTocallAdmin(request,pk):
    user1 = User.objects.get(id = pk)

    contacts = StudentContact.objects.filter(lead_follow_up = user1,next_follow_up__lt = date.today(),active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts),
    "user1":user1
    
    }  

    return render(request,"pendingcontacts.html",context)

@login_required(login_url="login")
def TodaysFollowUpAdmin(request,pk):
    user1 = User.objects.get(id = pk)

    contacts = StudentContact.objects.filter(lead_follow_up = user1,next_follow_up = date.today(),active = True).exclude(follow_up_status = "Not Called")
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts),
    "user1":user1
    
    }  

    return render(request,"todays_follow_up.html",context)


@login_required(login_url="login")
def MyAssignmentsAdmin(request,pk):
    user1 = User.objects.get(id = pk)

    contacts = StudentContact.objects.filter(lead_follow_up = user1,follow_up_status = "Not Called",active = True ).order_by("next_follow_up")
    contacts_count = StudentContact.objects.filter(lead_follow_up = user1,follow_up_status = "Not Called",active = True).count()
    users = User.objects.all()
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)

    context = {
        "contacts":page_obj,
        "users":users,
        "contacts_count":contacts_count,
        "user1":user1
    }  
    return render(request,"myassignments.html",context)

@login_required(login_url="login")
def UpcommingFollowUpAdmin(request,pk):
    user1 = User.objects.get(id = pk)
    contacts = StudentContact.objects.filter(lead_follow_up = user1,next_follow_up__gt = date.today(),active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts),
    "user1":user1
    }  

    return render(request,"upcommingfollowup.html",context)

@login_required(login_url="login")
def CompletedTodayAdmin(request,pk):
    user1 = User.objects.get(id = pk)
    contacts = StudentContact.objects.filter(lead_follow_up = user1,last_follow_up = date.today(),active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts),
    "user1":user1
    }  

    return render(request,"leadscompletedtoday.html",context)

def CompletedYesterday(request,pk):
    this_month = timezone.now().month
    today = timezone.now()
    start_date = today + timedelta(days=-1)
    user1 = User.objects.get(id = pk)

    contacts = StudentContact.objects.filter(lead_follow_up = user1,last_follow_up = start_date )

    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    
    context = {
        "contacts":page_obj,
        "contacts_count":len(contacts),
        "user1":user1,
        "yesterday":start_date
    }
    return render(request,"yesterdayscompletion.html",context)



def AllCallsAdminview(request,pk):
    user1 = User.objects.get(id = pk)
    contacts = StudentContact.objects.filter(lead_follow_up = user1)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    
    context = {
        "contacts":page_obj,
        "contacts_count":len(contacts),
        "user1":user1,

    }
    return render(request,"allcallsbyuseradminview.html",context)

def RejectedAdminview(request,pk):
    user1 = User.objects.get(id = pk)
    contacts = StudentContact.objects.filter(lead_follow_up = user1, active=False)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    
    context = {
        "contacts":page_obj,
        "contacts_count":len(contacts),
        "user1":user1,

    }
    return render(request,"rejectedcalladminview.html",context)







# report generation for data 




this_month = timezone.now().month
today = timezone.now()
start_date = today + timedelta(days=-5)
end_date = today + timedelta(days=5)
resign_date = today +timedelta(days = -30)

@login_required(login_url='SignIn')
def FullDataReport(request):
    
    date = timezone.now().month
    date_year = timezone.now().year
    response = HttpResponse(content_type = 'text/csv')
    response['Content-Disposition'] = 'attachment; filename=Leaddatafullreport{}-{}.csv'.format(date,date_year)
    
    STUDENT = StudentContact.objects.all().order_by("added_date")
    def generate_serial_number():
        current_time = datetime.now()
        serial_number = current_time.strftime("%Y%m%d%H%M%S")
        return serial_number
    TokenU = generate_serial_number()
    writer = csv.writer(response)
    response.write('\n')  # Move to the next line after the first row
    response.write("FULL STUDENT DATABASE ")  # Write the unique report number to the next line
    writer.writerow(["Sl No",'Name',"PHONE NUMBER","FOLLOWUP DATE","REMARKS","COLLEGE","COURSE","NO: FOLLOW UP","EMAIL","FOLLOW UP BY","FOLLOWUP STATUS","NEXT FOLLOW UP"])
    counter = 0
    for i in STUDENT:
        counter +=1
        try:
            staff = i.lead_follow_up.first_name
        except:
            staff = "Not assigned"
        writer.writerow([counter,i.name,i.contact_number,i.last_follow_up,i.last_status,i.collage,i.study_streem,i.number_follow_up,i.email,staff,i.follow_up_status,i.next_follow_up])
    response.write('\n')  # Move to the next line after the first row
    response.write(f"Doc Number: {TokenU}")  # Write the unique report number to the next line
    return response

from datetime import date as dt

@login_required(login_url='SignIn')
def MyReportTodaysFollowUp(request):
    date = timezone.now().month
    date_year = timezone.now().year
    response = HttpResponse(content_type = 'text/csv')
    response['Content-Disposition'] = 'attachment; filename=Callreporton-{}-{}.csv'.format(request.user.first_name,dt.today())
    
    STUDENT = StudentContact.objects.filter(lead_follow_up = request.user,last_follow_up = dt.today(),active = True)
    def generate_serial_number():
        current_time = datetime.now()
        serial_number = current_time.strftime("%Y%m%d%H%M%S")
        return serial_number
    TokenU = generate_serial_number()
    writer = csv.writer(response)
    response.write('\n')  # Move to the next line after the first row
    response.write("FULL STUDENT DATABASE ")  # Write the unique report number to the next line
    writer.writerow(["Sl No",'Name',"PHONE NUMBER","FOLLOWUP DATE","REMARKS","COLLEGE","COURSE","NO: FOLLOW UP","EMAIL","FOLLOW UP BY","FOLLOWUP STATUS","NEXT FOLLOW UP"])
    counter = 0
    for i in STUDENT:
        counter +=1
        try:
            staff = i.lead_follow_up.first_name
        except:
            staff = "Not assigned"
        writer.writerow([counter,i.name,i.contact_number,i.last_follow_up,i.last_status,i.collage,i.study_streem,i.number_follow_up,i.email,staff,i.follow_up_status,i.next_follow_up])
    response.write('\n')  # Move to the next line after the first row
    response.write(f"Doc Number: {TokenU}")  # Write the unique report number to the next line
    return response


@login_required(login_url='SignIn')
def UpdatedDataAll(request):
    date = timezone.now().month
    date_year = timezone.now().year
    response = HttpResponse(content_type = 'text/csv')
    response['Content-Disposition'] = 'attachment; filename=Callreporton-{}-{}.csv'.format("fulldata Report on",dt.today())
    
    STUDENT = StudentContact.objects.filter(active = True).exclude(follow_up_status = "Not Called")
    def generate_serial_number():
        current_time = datetime.now()
        serial_number = current_time.strftime("%Y%m%d%H%M%S")
        return serial_number
    TokenU = generate_serial_number()
    writer = csv.writer(response)
    response.write('\n')  # Move to the next line after the first row
    response.write("FULL STUDENT DATABASE ")  # Write the unique report number to the next line
    writer.writerow(["Sl No",'Name',"PHONE NUMBER","FOLLOWUP DATE","REMARKS","COLLEGE","COURSE","NO: FOLLOW UP","EMAIL","FOLLOW UP BY","FOLLOWUP STATUS","NEXT FOLLOW UP"])
    counter = 0
    for i in STUDENT:
        counter +=1
        try:
            staff = i.lead_follow_up.first_name
        except:
            staff = "Not assigned"
        writer.writerow([counter,i.name,i.contact_number,i.last_follow_up,i.last_status,i.collage,i.study_streem,i.number_follow_up,i.email,staff,i.follow_up_status,i.next_follow_up])
    response.write('\n')  # Move to the next line after the first row
    response.write(f"Doc Number: {TokenU}")  # Write the unique report number to the next line
    return response

@login_required(login_url='SignIn')
def DeleteContacts(request):
    if request.method == "POST":
        selected_contacts = request.POST.getlist('contact_id')
    
        for item in selected_contacts:
            contact = StudentContact.objects.get(id = int(item))
            contact.delete()
        messages.info(request,"selected contact deleted.....")
        return redirect("AddContact")
    return redirect("AddContact")


def RejectedCallList(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,active = False)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts),
    "user1":request.user
    }  

    return render(request,"rejectedcalls.html",context)


def AllCallList(request):
    contacts = StudentContact.objects.filter(lead_follow_up = request.user,active = True)
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
    "contacts":page_obj,
    "contacts_count":len(contacts),
    "user1":request.user
    }  

    return render(request,"Allcalls.html",context)


from django.db.models import Q
@login_required(login_url="login")
def Search(request):
    contacts = None
    if request.method == "POST":
        search = request.POST['search']
        contacts = StudentContact.objects.filter(
            Q(name__icontains=search) | Q(collage__icontains=search)
        )
    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    
    context = {
        "contacts":page_obj
    }
    return render(request,"searchresults.html",context)


def SearchBydate(request):
    contacts = None
    if request.method == "POST":
        sdate = request.POST['sdate']
        edate = request.POST['edate']
        leadstatus = request.POST["leadstatus"]
        if leadstatus == "all":
            contacts = StudentContact.objects.filter(last_follow_up__gte = sdate, last_follow_up__lte = edate, active = True)
        else:
            contacts = StudentContact.objects.filter(last_follow_up__gte = sdate, last_follow_up__lte = edate, follow_up_status = leadstatus )



    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    
    context = {
        "contacts":page_obj
    }
    return render(request,"searchresults.html",context)



def SearchBydateadmin(request):
    contacts = None
    if request.method == "POST":
        sdate = request.POST['sdate']
        edate = request.POST['edate']
        leadstatus = request.POST["leadstatus"]
        user = request.POST['user']
        try:
            user = User.objects.get(id = int(user))
        except:
            user = "all"
        if leadstatus == "all":
            if user == "all":
                contacts = StudentContact.objects.filter(last_follow_up__gte = sdate, last_follow_up__lte = edate, active = True)
            else:
                contacts = StudentContact.objects.filter(lead_follow_up = user,last_follow_up__gte = sdate, last_follow_up__lte = edate, active = True,)

        else:
            if user == "all":
                contacts = StudentContact.objects.filter(last_follow_up__gte = sdate, last_follow_up__lte = edate, follow_up_status = leadstatus )
            else:
                contacts = StudentContact.objects.filter(lead_follow_up = user,last_follow_up__gte = sdate, last_follow_up__lte = edate, active = True,)

    p = Paginator(contacts, 30)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    
    context = {
        "contacts":page_obj
    }
    return render(request,"searchresults.html",context)









    
        